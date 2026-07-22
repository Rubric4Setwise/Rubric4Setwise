#!/usr/bin/env python3
"""
Rubric scoring script (LLM-as-judge).

Given a JSONL file where each entry contains:
  - question / answer / data_source
  - docs:            candidate document set (each item has original_id, doc, ...)
  - selected_docs:   dict {ranker_name: [selected original_id, ...]}
  - hybrid_rubrics:  9-dimension rubric (each dim is {"rubrics": [...], ...})

For every entry the script iterates over all configured rerankers, feeds the
ranker's selected doc set together with the 9-dimension rubric into a single
LLM call, and asks the judge for atomic (0-4) scores. dimension_score /
rubric_avg / overall are ALL computed in this script -- the LLM only supplies
per_doc / set_level atomic scores plus evidence/reason.

Key features (merged from the earlier v1/v2 scripts):
- 0-4 scale, one prompt scores 9 dimensions at once.
- RELEVANCE GATE: if every doc has Relevance=0, the other 8 dimensions are
  marked skipped (dimension_score=None) instead of being scored.
- Conflict "not applicable" when the doc set has fewer than two comparable
  statements about the same fact.
- Dimension-level resume: if an output already exists, entries with only some
  dimensions missing are repaired instead of being fully re-scored.
- Final de-duplicating rewrite of the output JSONL and an Excel summary.

Prompt template is kept in `scoring_prompt.txt` next to this script.

Usage:
    python scoring.py <input.jsonl> <output.jsonl> \\
        [--rankers r1,r2,...] [--rubric-field hybrid_rubrics] [--docs-field docs] \\
        [--selected-field selected_docs] [--max-docs N] \\
        [--app-id ID] [--app-key KEY] [--limit N] [--offset N] [--concurrency N] \\
        [--base-url URL] [--model NAME] [--prompt-file PATH]
"""

import argparse
import json
import os
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Optional

import requests

try:
    from tqdm import tqdm
except ImportError:
    print("[ERROR] tqdm is required. Install with: pip install tqdm")
    raise


# ============================================================
# Config
# ============================================================
DEFAULT_BASE_URL = "http://llm-api.model-eval.woa.com/v1/chat/completions"
DEFAULT_MODEL = "api_naci_default_deepseek-v4-pro"

MAX_RETRIES = 3
RETRY_DELAY_BASE = 2
TIMEOUT = 3600
# One call produces per-doc x per-dimension JSON -- give the LLM plenty of room.
MAX_TOKENS = 16384
# Number of parse retries per entry when the returned JSON cannot be parsed.
PARSE_MAX_RETRIES = 3

DEFAULT_APP_ID = os.environ.get("LLM_APP_ID", "")
DEFAULT_APP_KEY = os.environ.get("LLM_APP_KEY", "")

DEFAULT_PROMPT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                   "scoring_prompt.txt")

# Rankers scored per entry (in this order); only rankers present in
# selected_docs are actually scored.
RANKERS = [
    "bge-reranker-large", "rankllama", "rankvicuna", "rankzephyr",
    "monot5", "rankt5", "rank4gen", "setr", "reasonrank-7b", "rearank-7b",
    "setwise-sft-7b", "rank1-7b", "bm25-baseline",
    # Rubric4Setwise (our method)
    "rubric4setwise",
]

# ranker -> (display_name, size) for the Excel summary
RANKER_META = {
    "bm25-baseline":      ("Only Retrieval", "-"),
    "bge-reranker-large": ("bge-reranker-large", "550M"),
    "monot5":             ("MonoT5(3B)", "3B"),
    "rankt5":             ("RankT5(3B)", "3B"),
    "rankllama":          ("RankLlama(7B)", "7B"),
    "rankvicuna":         ("RankVicuna(7B)", "7B"),
    "rankzephyr":         ("RankZephyr(7B)", "7B"),
    "rank4gen":           ("Rank4Gen", "-"),
    "rank1-7b":           ("Rank1(7B)", "7B"),
    "rearank-7b":         ("Rearank(7B)", "7B"),
    "reasonrank-7b":      ("ReasonRank(7B)", "7B"),
    "setwise-sft-7b":     ("Setwise(7B)", "7B"),
    "setr":               ("SetR(8B)", "8B"),
    "rubric4setwise":     ("Rubric4Setwise", "8B"),
}

# Excel row grouping (section title, ordered rankers).
RANKER_SECTIONS = [
    (None, ["bm25-baseline", "bge-reranker-large", "monot5", "rankt5",
            "rankllama", "rankvicuna", "rankzephyr"]),
    ("Reasoning-Enhanced Reranking", ["rank1-7b", "rearank-7b", "reasonrank-7b"]),
    ("Setwise Reranking", ["setwise-sft-7b", "setr", "rank4gen"]),
    ("Ours", ["rubric4setwise"]),
]


# ============================================================
# Dimension definitions
# ============================================================
LEVELS = {
    "Doc-Level":    ["Relevance", "Authenticity", "Quality"],
    "Set-Level":    ["Complementarity", "Redundancy", "Conflict"],
    "Global-Level": ["Completeness", "Density", "Reachability"],
}
DOC_DIMS = set(LEVELS["Doc-Level"])
ALL_DIMS = [d for dims in LEVELS.values() for d in dims]
DIM_TO_LEVEL = {d: lv for lv, dims in LEVELS.items() for d in dims}

VALID_SCORES = {0, 1, 2, 3, 4}


def is_doc_level(dim: str) -> bool:
    return dim in DOC_DIMS


# ============================================================
# Prompt construction
# ============================================================
def load_prompt_template(path: str) -> str:
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def select_docs_for_ranker(item: dict, ranker: str,
                           docs_field: str, selected_field: str) -> List[dict]:
    """Return the doc list that `ranker` picked, in the order recorded
    under `item[selected_field][ranker]` (a list of original_id values)."""
    docs = item.get(docs_field) or []
    id2doc = {}
    for d in docs:
        if isinstance(d, dict) and d.get("original_id") is not None:
            id2doc[int(d["original_id"])] = d

    selected = item.get(selected_field) or {}
    if not isinstance(selected, dict):
        return []
    sel_ids = selected.get(ranker)
    if not isinstance(sel_ids, list):
        return []

    out = []
    for oid in sel_ids:
        try:
            oid_int = int(oid)
        except (TypeError, ValueError):
            continue
        d = id2doc.get(oid_int)
        if d is not None:
            out.append(d)
    return out


def format_docs(docs: List[dict], max_docs: Optional[int]):
    """Render the doc set into prompt text. Each line looks like
    `[original_id] <text>`; the LLM must reuse the bracketed ids in per_doc.
    Returns (rendered_text, list_of_docs_actually_used)."""
    if max_docs is not None:
        docs = docs[:max_docs]
    lines = []
    for i, d in enumerate(docs, start=1):
        oid = d.get("original_id", i)
        text = str(d.get("doc", "")).strip()
        lines.append(f"[{oid}] {text}")
    return "\n\n".join(lines), docs


def get_rubric_list(rubric_block: dict, dim: str) -> List[str]:
    """Pull the list of rubric strings for one dimension out of `hybrid_rubrics`."""
    entry = rubric_block.get(dim)
    if isinstance(entry, dict):
        val = entry.get("rubrics", [])
    elif isinstance(entry, list):
        val = entry
    else:
        val = []
    if isinstance(val, str):
        val = [val]
    if not isinstance(val, list):
        return []
    return [str(x).strip() for x in val if str(x).strip()]


def format_rubrics(rubric_block: dict):
    """Render 9 rubric dimensions into prompt text.
    Returns (rendered_text, {dim: [rubric_str, ...]})."""
    dim_rubrics: Dict[str, List[str]] = {}
    lines = []
    for level, dims in LEVELS.items():
        for dim in dims:
            rubrics = get_rubric_list(rubric_block, dim)
            dim_rubrics[dim] = rubrics
            if not rubrics:
                continue
            gran = ("score each document (per_doc)" if is_doc_level(dim)
                    else "one overall score for the whole doc set (set_level)")
            lines.append(f"[{dim}] ({level}, {gran})")
            for j, r in enumerate(rubrics, start=1):
                lines.append(f"  {dim} #{j}: {r}")
            lines.append("")
    return "\n".join(lines).strip(), dim_rubrics


def build_prompt(prompt_template: str, docs_str: str, rubrics_str: str) -> str:
    """Fill `{docs_str}` / `{rubrics_str}` in the external prompt template."""
    return prompt_template.format(docs_str=docs_str, rubrics_str=rubrics_str)


# ============================================================
# LLM call
# ============================================================
def get_headers(app_id: str, app_key: str) -> dict:
    return {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {app_id}:{app_key}",
    }


def call_llm(prompt: str, app_id: str, app_key: str, base_url: str, model: str):
    """Call the judge LLM with retry. Returns (content, finish_reason).
    finish_reason == "length" means the output was cut off by max_tokens."""
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": MAX_TOKENS,
        "stream": False,
    }

    last_error = None
    for attempt in range(MAX_RETRIES):
        try:
            resp = requests.post(base_url, headers=get_headers(app_id, app_key),
                                 json=payload, timeout=TIMEOUT)
            if resp.status_code != 200:
                last_error = f"HTTP {resp.status_code}: {resp.text[:300]}"
                tqdm.write(f"  [WARN] LLM request failed (attempt {attempt+1}/{MAX_RETRIES}): {last_error}")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY_BASE ** (attempt + 1))
                continue

            data = resp.json()
            content, finish_reason = None, None
            if "choices" in data and data["choices"]:
                content = data["choices"][0].get("message", {}).get("content")
                finish_reason = data["choices"][0].get("finish_reason")
            if content is None:
                content = json.dumps(data, ensure_ascii=False)
                tqdm.write(f"  [WARN] Unexpected response payload: {content[:200]}")
            return content, finish_reason

        except requests.exceptions.Timeout:
            last_error = "timeout"
            tqdm.write(f"  [WARN] LLM request timeout (attempt {attempt+1}/{MAX_RETRIES})")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY_BASE ** (attempt + 1))
        except requests.exceptions.ConnectionError as e:
            last_error = f"connection error: {e}"
            tqdm.write(f"  [WARN] LLM connection failed (attempt {attempt+1}/{MAX_RETRIES})")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY_BASE ** (attempt + 1))
        except (json.JSONDecodeError, KeyError, TypeError, IndexError) as e:
            last_error = f"response parse failed: {e}"
            tqdm.write(f"  [WARN] LLM response parse failed (attempt {attempt+1}/{MAX_RETRIES}): {e}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY_BASE ** (attempt + 1))

    raise RuntimeError(f"LLM failed after {MAX_RETRIES} retries: {last_error}")


# ============================================================
# Parsing / score computation
# ============================================================
def _to_score(x) -> Optional[int]:
    """Coerce a value into an integer 0-4 score; return None if invalid."""
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        v = int(round(x))
    else:
        m = re.search(r"[0-4]", str(x))
        if not m:
            return None
        v = int(m.group(0))
    return v if v in VALID_SCORES else None


def parse_score_obj(raw_text: str) -> Optional[Dict]:
    """Parse the LLM output into a `{dim: {...}}` score dict; None on failure."""
    text = (raw_text or "").strip()

    fence = re.search(r"```(?:json)?\s*(.*?)\s*```", text, re.DOTALL)
    if fence:
        text = fence.group(1).strip()

    obj = re.search(r"\{.*\}", text, re.DOTALL)
    if obj:
        text = obj.group(0)

    try:
        data = json.loads(text)
    except (json.JSONDecodeError, TypeError):
        return None

    if not isinstance(data, dict):
        return None

    # Accept both {"score": {...}} and a bare {...} payload.
    score = data.get("score") if isinstance(data.get("score"), dict) else data
    if not isinstance(score, dict):
        return None
    return score


def round1(x: Optional[float]) -> Optional[float]:
    return None if x is None else round(x + 1e-9, 1)


def compute_scores(score_obj: dict, dim_rubrics: Dict[str, List[str]]) -> Dict:
    """Compute per-rubric rubric_avg and per-dimension dimension_score from
    the LLM-provided atomic scores. `dim_rubrics` is the set of rubrics we
    actually sent to the LLM (used for alignment / fallbacks)."""
    result: Dict[str, Dict] = {}
    for dim in ALL_DIMS:
        sent = dim_rubrics.get(dim, [])
        level = "doc" if is_doc_level(dim) else "set"
        if not sent:
            # No rubric was sent for this dimension -- nothing to score.
            result[dim] = {"level": level, "rubrics": [], "dimension_score": None}
            continue

        dim_entry = score_obj.get(dim) if isinstance(score_obj, dict) else None
        returned = []
        if isinstance(dim_entry, dict):
            returned = dim_entry.get("rubrics", []) or []
        elif isinstance(dim_entry, list):
            returned = dim_entry

        rubrics_out = []
        rubric_avgs = []
        for i, rubric_text in enumerate(sent):
            r_ret = returned[i] if i < len(returned) and isinstance(returned[i], dict) else {}

            if level == "doc":
                per_doc_in = r_ret.get("per_doc", []) if isinstance(r_ret, dict) else []
                per_doc_out, valid = [], []
                if isinstance(per_doc_in, list):
                    for pd in per_doc_in:
                        if not isinstance(pd, dict):
                            continue
                        sc = _to_score(pd.get("score"))
                        per_doc_out.append({
                            "doc_id": pd.get("doc_id"),
                            "evidence": pd.get("evidence", ""),
                            "score": sc,
                            "reason": pd.get("reason", ""),
                        })
                        if sc is not None:
                            valid.append(sc)
                rubric_avg = round1(sum(valid) / len(valid)) if valid else None
                rubrics_out.append({
                    "rubric": rubric_text,
                    "per_doc": per_doc_out,
                    "rubric_avg": rubric_avg,
                })
            else:
                sl_in = r_ret.get("set_level", {}) if isinstance(r_ret, dict) else {}
                sc, evidence, reason = None, "", ""
                if isinstance(sl_in, dict):
                    sc = _to_score(sl_in.get("score"))
                    evidence = sl_in.get("evidence", "")
                    reason = sl_in.get("reason", "")
                # Fallback: some models put `score` directly at the rubric level.
                if sc is None:
                    sc = _to_score(r_ret.get("score")) if isinstance(r_ret, dict) else None
                rubric_avg = round1(float(sc)) if sc is not None else None
                rubrics_out.append({
                    "rubric": rubric_text,
                    "set_level": {"evidence": evidence, "score": sc, "reason": reason},
                    "rubric_avg": rubric_avg,
                })

            if rubric_avg is not None:
                rubric_avgs.append(rubric_avg)

        dimension_score = round1(sum(rubric_avgs) / len(rubric_avgs)) if rubric_avgs else None
        result[dim] = {"level": level, "rubrics": rubrics_out, "dimension_score": dimension_score}

    # ---- Conflict "not applicable" marker ------------------------------------
    # When the LLM outputs Conflict as an explicit JSON null, we treat it as
    # "not applicable" (fewer than two comparable statements). This is
    # different from a missing key -- an explicit null is a decision, not a
    # scoring failure, so the resume path should not retry it.
    if dim_rubrics.get("Conflict"):
        conf = result.get("Conflict")
        explicit_null = (isinstance(score_obj, dict) and ("Conflict" in score_obj)
                         and score_obj.get("Conflict") is None)
        if isinstance(conf, dict) and conf.get("dimension_score") is None and explicit_null:
            conf["rubrics"] = []
            conf["not_applicable"] = ("fewer than two comparable statements about the same "
                                      "fact; Conflict not applicable")

    # ---- Relevance gate ------------------------------------------------------
    _apply_relevance_gate(result)
    return result


def _apply_relevance_gate(result: Dict[str, Dict]) -> bool:
    """If every per_doc Relevance score is 0, mark the other 8 dimensions as
    skipped (dimension_score=None). Returns whether the gate was triggered."""
    rel = result.get("Relevance") or {}
    rel_scores = []
    for rub in rel.get("rubrics", []) or []:
        for pd in rub.get("per_doc", []) or []:
            s = pd.get("score")
            if s is not None:
                rel_scores.append(s)
    # Only trigger when at least one Relevance score exists and all are 0.
    if not rel_scores or max(rel_scores) > 0:
        return False
    for dim in ALL_DIMS:
        if dim == "Relevance":
            continue
        level = "doc" if is_doc_level(dim) else "set"
        result[dim] = {"level": level, "rubrics": [], "dimension_score": None,
                       "skipped": "all docs irrelevant (Relevance=0); dimension not evaluated"}
    return True


def compute_overall(scores: Dict[str, Dict]) -> Dict:
    """Aggregate dimension_score by level; produce per-level means + all-dims mean."""
    overall, all_dim_scores = {}, []
    for level, dims in LEVELS.items():
        vals = [scores[d]["dimension_score"] for d in dims
                if scores.get(d, {}).get("dimension_score") is not None]
        overall[level] = round1(sum(vals) / len(vals)) if vals else None
        all_dim_scores.extend(vals)
    overall["all_dims_mean"] = round1(sum(all_dim_scores) / len(all_dim_scores)) if all_dim_scores else None
    return overall


def refresh_entry_overall(entry: dict) -> dict:
    """Recompute overall for each ranker block based on its current scores.
    Used right before the final rewrite so overall always matches the data."""
    br = entry.get("rubric_scores_by_ranker")
    if isinstance(br, dict):
        for _rk, blk in br.items():
            if isinstance(blk, dict) and isinstance(blk.get("rubric_scores"), dict):
                blk["overall"] = compute_overall(blk["rubric_scores"])
    return entry


# ============================================================
# I/O helpers
# ============================================================
def load_jsonl(path: str) -> List[dict]:
    out = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                out.append(json.loads(line))
    return out


def make_key(entry: dict):
    return (entry.get("question"), entry.get("answer"), entry.get("data_source"))


# ============================================================
# Resume: detect missing dimensions + partial repair
# ============================================================
def block_missing_dims(block: dict, dim_rubrics: Dict[str, List[str]]) -> List[str]:
    """List dimensions that SHOULD be scored but currently are not.

    Cases considered as "already done" (NOT missing):
      - the dimension had no rubric in the input;
      - the relevance gate was triggered for this block ("skipped" marker);
      - Conflict is marked not_applicable;
      - a valid dimension_score already exists.
    """
    missing = []
    scores = block.get("rubric_scores") if isinstance(block, dict) else None
    if not isinstance(scores, dict):
        return [d for d in ALL_DIMS if dim_rubrics.get(d)]
    gated = bool(block.get("low_relevance_skip"))
    for dim in ALL_DIMS:
        if not dim_rubrics.get(dim):
            continue
        if gated and dim != "Relevance":
            continue  # intentionally None under the gate
        s = scores.get(dim)
        if not isinstance(s, dict):
            missing.append(dim)
            continue
        if s.get("skipped") or s.get("not_applicable"):
            continue
        if s.get("dimension_score") is None:
            missing.append(dim)
    return missing


def ranker_work(block: dict, dim_rubrics: Dict[str, List[str]]):
    """Decide what to do with an existing ranker block:
      ('done', None)              -- nothing to do
      ('full', None)              -- rerun the whole block
      ('repair', [missing_dims])  -- only re-score the missing dimensions
    """
    if not isinstance(block, dict) or block.get("scoring_error"):
        return ("full", None)
    if not isinstance(block.get("rubric_scores"), dict):
        return ("full", None)
    missing = block_missing_dims(block, dim_rubrics)
    if not missing:
        return ("done", None)
    # Relevance gate depends on Relevance, so a missing Relevance forces full rerun.
    if "Relevance" in missing:
        return ("full", None)
    return ("repair", missing)


# ============================================================
# Core scoring
# ============================================================
def score_one_ranker(idx: int, total: int, item: dict, ranker: str,
                     dim_rubrics: Dict[str, List[str]], rubrics_str: str,
                     docs_field: str, selected_field: str, max_docs: Optional[int],
                     prompt_template: str, app_id: str, app_key: str,
                     base_url: str, model: str) -> dict:
    """Score every dimension for one (entry, ranker) pair (full scoring)."""
    sel_docs = select_docs_for_ranker(item, ranker, docs_field, selected_field)
    docs_str, used_docs = format_docs(sel_docs, max_docs)

    block = {"scored_doc_ids": [d.get("original_id") for d in used_docs]}

    if not used_docs:
        block["rubric_scores"] = compute_scores({}, dim_rubrics)
        block["overall"] = compute_overall(block["rubric_scores"])
        block["scoring_error"] = f"ranker={ranker} selected no docs"
        tqdm.write(f"  [{idx+1}/{total}] [{ranker}] skipped (no selected docs)")
        return block

    prompt = build_prompt(prompt_template, docs_str, rubrics_str)

    parsed, last_raw, last_finish = None, "", None
    for p_attempt in range(PARSE_MAX_RETRIES):
        raw, finish_reason = call_llm(prompt, app_id, app_key, base_url, model)
        last_raw, last_finish = raw, finish_reason
        parsed = parse_score_obj(raw)
        if parsed is not None:
            break
        trunc = " (finish_reason=length, likely truncated by max_tokens)" if finish_reason == "length" else ""
        tqdm.write(f"  [WARN] [{idx+1}] [{ranker}] parse failed "
                   f"(attempt {p_attempt+1}/{PARSE_MAX_RETRIES}){trunc}")
        if p_attempt < PARSE_MAX_RETRIES - 1:
            time.sleep(1)

    if parsed is None:
        tqdm.write(f"  [ERROR] [{idx+1}] [{ranker}] parse still failed after "
                   f"{PARSE_MAX_RETRIES} attempts (finish_reason={last_finish}); raw output:")
        tqdm.write(f"  ----- RAW BEGIN ({ranker}) -----\n{last_raw}\n  ----- RAW END -----")
        block["rubric_scores"] = compute_scores({}, dim_rubrics)
        block["overall"] = compute_overall(block["rubric_scores"])
        block["scoring_error"] = f"parse failed (finish_reason={last_finish})"
        block["scoring_raw"] = last_raw
        return block

    scores = compute_scores(parsed, dim_rubrics)
    block["rubric_scores"] = scores
    block["overall"] = compute_overall(scores)
    gated = any(isinstance(scores.get(d), dict) and scores[d].get("skipped")
                for d in ALL_DIMS if d != "Relevance")
    if gated:
        block["low_relevance_skip"] = True
    tqdm.write(f"  [{idx+1}/{total}] [{ranker}] docs={len(used_docs)} "
               f"overall={block['overall']}"
               + ("  [low relevance: only Relevance counted]" if gated else ""))
    return block


def repair_one_ranker(idx: int, total: int, item: dict, ranker: str,
                      existing_block: dict, rubric_block: dict, missing_dims: List[str],
                      docs_field: str, selected_field: str, max_docs: Optional[int],
                      prompt_template: str, app_id: str, app_key: str,
                      base_url: str, model: str) -> dict:
    """Re-score only the missing dimensions and merge them into `existing_block`."""
    sel_docs = select_docs_for_ranker(item, ranker, docs_field, selected_field)
    docs_str, used_docs = format_docs(sel_docs, max_docs)
    if not used_docs:
        return existing_block  # nothing to score, keep as-is

    sub_block = {d: rubric_block.get(d, {}) for d in missing_dims if d in rubric_block}
    rubrics_str, sub_dim_rubrics = format_rubrics(sub_block)

    prompt = build_prompt(prompt_template, docs_str, rubrics_str)
    parsed, last_raw, last_finish = None, "", None
    for p_attempt in range(PARSE_MAX_RETRIES):
        raw, finish_reason = call_llm(prompt, app_id, app_key, base_url, model)
        last_raw, last_finish = raw, finish_reason
        parsed = parse_score_obj(raw)
        if parsed is not None:
            break
        if p_attempt < PARSE_MAX_RETRIES - 1:
            time.sleep(1)

    if parsed is None:
        tqdm.write(f"  [WARN] [{idx+1}] [{ranker}] dimension repair parse failed; "
                   f"leaving {missing_dims} for the next run")
        return existing_block

    new_scores = compute_scores(parsed, sub_dim_rubrics)
    merged = dict(existing_block)
    rs = dict(merged.get("rubric_scores") or {})
    for d in missing_dims:
        if isinstance(new_scores.get(d), dict):
            rs[d] = new_scores[d]
    merged["rubric_scores"] = rs
    merged["overall"] = compute_overall(rs)
    gated = any(isinstance(rs.get(d), dict) and rs[d].get("skipped")
                for d in ALL_DIMS if d != "Relevance")
    if gated:
        merged["low_relevance_skip"] = True
    elif "low_relevance_skip" in merged:
        del merged["low_relevance_skip"]
    merged.pop("scoring_error", None)
    merged.pop("scoring_raw", None)
    filled = [d for d in missing_dims if isinstance(rs.get(d), dict)
              and (rs[d].get("dimension_score") is not None or rs[d].get("not_applicable"))]
    tqdm.write(f"  [{idx+1}/{total}] [{ranker}] repaired {filled} overall={merged['overall']}")
    return merged


# ============================================================
# Excel summary
# ============================================================
def _collect_stats(output_file: str):
    """Return (accum, passages, n_by_ranker, n_lines) from the output JSONL.

    accum[ranker][dim]  = [dimension_score(0-4), ...]
    passages[ranker]    = [num scored docs per entry, ...]
    n_by_ranker[ranker] = number of entries that produced at least one score
    """
    accum: Dict[str, Dict[str, List[float]]] = {}
    passages: Dict[str, List[int]] = {}
    n_by_ranker: Dict[str, int] = {}
    n_lines = 0

    if os.path.exists(output_file):
        with open(output_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    entry = json.loads(line)
                except json.JSONDecodeError:
                    continue
                br = entry.get("rubric_scores_by_ranker")
                if not isinstance(br, dict):
                    continue
                n_lines += 1
                for ranker, block in br.items():
                    if not isinstance(block, dict):
                        continue
                    scores = block.get("rubric_scores") or {}
                    dim_accum = accum.setdefault(ranker, {})
                    counted = False
                    for dim in ALL_DIMS:
                        ds = scores.get(dim, {})
                        val = ds.get("dimension_score") if isinstance(ds, dict) else None
                        if val is not None:
                            dim_accum.setdefault(dim, []).append(float(val))
                            counted = True
                    doc_ids = block.get("scored_doc_ids") or []
                    passages.setdefault(ranker, []).append(len(doc_ids))
                    if counted:
                        n_by_ranker[ranker] = n_by_ranker.get(ranker, 0) + 1
    return accum, passages, n_by_ranker, n_lines


def write_stats(output_file: str, rankers: List[str]) -> str:
    """Aggregate the output JSONL into an Excel workbook next to the output
    file. Raw 0-4 scores are rendered as percentages (x25, two decimals)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    accum, passages, _, n_lines = _collect_stats(output_file)

    def mean(vals):
        return sum(vals) / len(vals) if vals else None

    def pct(v):
        return round(v * 25, 2) if isinstance(v, (int, float)) else None

    fill_doc = PatternFill("solid", fgColor="FCE4EC")   # pink
    fill_set = PatternFill("solid", fgColor="FFF2CC")   # amber
    fill_glb = PatternFill("solid", fgColor="DEEBF7")   # blue
    fill_ovr = PatternFill("solid", fgColor="E2EFDA")   # green
    fill_sec = PatternFill("solid", fgColor="F2F2F2")   # section separator (gray)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    bold = Font(bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "rubric_scores"

    # Column layout:
    # A Ranker | B Size | C done | D #Passages
    # E-H Doc-Level(Relevance,Authenticity,Quality,Avg)
    # I-L Set-Level(Complementarity,Redundancy,Conflict,Avg)
    # M-P Global-Level(Completeness,Density,Reachability,Avg)
    # Q Overall
    headers_sub = ["Ranker", "Size", "done", "# of Passages",
                   "Relevance", "Authenticity", "Quality", "Avg",
                   "Complementarity", "Redundancy", "Conflict", "Avg",
                   "Completeness", "Density", "Reachability", "Avg",
                   "Overall"]
    ncol = len(headers_sub)

    # Row 1: group headers (merged); row 2: sub-headers under groups.
    ws.merge_cells("A1:A2"); ws["A1"] = "Ranker"
    ws.merge_cells("B1:B2"); ws["B1"] = "Size"
    ws.merge_cells("C1:C2"); ws["C1"] = "done"
    ws.merge_cells("D1:D2"); ws["D1"] = "# of Passages"
    ws.merge_cells("E1:H1"); ws["E1"] = "Doc-Level"
    ws.merge_cells("I1:L1"); ws["I1"] = "Set-Level"
    ws.merge_cells("M1:P1"); ws["M1"] = "Global-Level"
    ws.merge_cells("Q1:Q2"); ws["Q1"] = "Overall"

    for ci, name in enumerate(headers_sub, start=1):
        if 5 <= ci <= 16:
            ws.cell(row=2, column=ci, value=name)

    group_fill = {**{c: fill_doc for c in "EFGH"},
                  **{c: fill_set for c in "IJKL"},
                  **{c: fill_glb for c in "MNOP"},
                  "Q": fill_ovr}
    for c in "ABCDEFGHIJKLMNOPQ":
        for r in (1, 2):
            cell = ws[f"{c}{r}"]
            cell.font = bold
            cell.alignment = center
            cell.border = border
            if c in group_fill:
                cell.fill = group_fill[c]

    row = 3

    def write_ranker_row(r: int, ranker: str):
        disp, size = RANKER_META.get(ranker, (ranker, "-"))
        dim_accum = accum.get(ranker, {})
        dim_means = {d: mean(dim_accum.get(d, [])) for d in ALL_DIMS}

        def level_avg(dims):
            vals = [dim_means[d] for d in dims if dim_means[d] is not None]
            return mean(vals) if vals else None

        doc_avg = level_avg(LEVELS["Doc-Level"])
        set_avg = level_avg(LEVELS["Set-Level"])
        glb_avg = level_avg(LEVELS["Global-Level"])
        all_vals = [dim_means[d] for d in ALL_DIMS if dim_means[d] is not None]
        overall = mean(all_vals) if all_vals else None
        n_pass = passages.get(ranker, [])
        avg_pass = round(mean(n_pass), 1) if n_pass else None
        has_data = ranker in accum

        vals = [
            disp, size, ("Y" if has_data else ""), avg_pass,
            pct(dim_means["Relevance"]), pct(dim_means["Authenticity"]), pct(dim_means["Quality"]), pct(doc_avg),
            pct(dim_means["Complementarity"]), pct(dim_means["Redundancy"]), pct(dim_means["Conflict"]), pct(set_avg),
            pct(dim_means["Completeness"]), pct(dim_means["Density"]), pct(dim_means["Reachability"]), pct(glb_avg),
            pct(overall),
        ]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.border = border
            cell.alignment = left if ci == 1 else center
            letter = get_column_letter(ci)
            if letter in group_fill:
                cell.fill = group_fill[letter]
        for letter in ["H", "L", "P", "Q"]:
            ws[f"{letter}{r}"].font = bold

    for section_title, section_rankers in RANKER_SECTIONS:
        if section_title is not None:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
            c = ws.cell(row=row, column=1, value=section_title)
            c.font = bold
            c.alignment = center
            c.fill = fill_sec
            for ci in range(1, ncol + 1):
                ws.cell(row=row, column=ci).border = border
            row += 1
        for ranker in section_rankers:
            write_ranker_row(row, ranker)
            row += 1

    # Any ranker that has data but is not in an explicit section: append at the bottom.
    listed = {rk for _, rks in RANKER_SECTIONS for rk in rks}
    extras = [rk for rk in accum if rk not in listed]
    for ranker in extras:
        write_ranker_row(row, ranker)
        row += 1

    widths = {"A": 20, "B": 8, "C": 6, "D": 13}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w
    for c in "EFGHIJKLMNOPQ":
        ws.column_dimensions[c].width = 13
    ws.freeze_panes = "A3"

    note_row = row + 1
    ws.cell(row=note_row, column=1,
            value=f"Note: percentage-scale scores (raw 0-4 x25), two decimals; entries counted = {n_lines}"
            ).font = Font(italic=True, size=9)

    base, _ = os.path.splitext(output_file)
    stats_path = base + "_stats.xlsx"
    wb.save(stats_path)
    return stats_path


# ============================================================
# Main
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Rubric scoring (LLM-as-judge)")
    parser.add_argument("input_file",  type=str, help="Input JSONL (with docs + rubric fields)")
    parser.add_argument("output_file", type=str, help="Output JSONL path")
    parser.add_argument("--rankers", type=str, default=None,
                        help="Comma-separated ranker list; defaults to the built-in list")
    parser.add_argument("--rubric-field",   type=str, default="hybrid_rubrics")
    parser.add_argument("--docs-field",     type=str, default="docs")
    parser.add_argument("--selected-field", type=str, default="selected_docs")
    parser.add_argument("--max-docs", type=int, default=None,
                        help="Score only the first N docs per ranker (default: all)")
    parser.add_argument("--app-id",  type=str, default=DEFAULT_APP_ID)
    parser.add_argument("--app-key", type=str, default=DEFAULT_APP_KEY)
    parser.add_argument("--base-url", type=str, default=DEFAULT_BASE_URL)
    parser.add_argument("--model",    type=str, default=DEFAULT_MODEL)
    parser.add_argument("--prompt-file", type=str, default=DEFAULT_PROMPT_FILE,
                        help="Path to the prompt template (default: scoring_prompt.txt)")
    parser.add_argument("--limit",  type=int, default=None, help="Max entries to process")
    parser.add_argument("--offset", type=int, default=0,    help="Start offset into the input file")
    parser.add_argument("--concurrency", type=int, default=5, help="Worker threads (default 5)")
    args = parser.parse_args()

    if not args.app_id or not args.app_key:
        raise SystemExit("[ERROR] Missing --app-id / --app-key (or LLM_APP_ID / LLM_APP_KEY env vars).")

    prompt_template = load_prompt_template(args.prompt_file)

    rankers = ([r.strip() for r in args.rankers.split(",") if r.strip()]
               if args.rankers else list(RANKERS))

    print(f"Model:    {args.model}")
    print(f"Base URL: {args.base_url}")
    print(f"Input:    {args.input_file}")
    print(f"Rankers ({len(rankers)}): {rankers}")
    print(f"Fields: rubric={args.rubric_field} | docs={args.docs_field} | "
          f"selected={args.selected_field} | max_docs={args.max_docs}")
    print(f"Prompt template: {args.prompt_file}")

    items = load_jsonl(args.input_file)
    print(f"Loaded {len(items)} entries")

    start = args.offset
    end = len(items) if args.limit is None else min(start + args.limit, len(items))
    items_to_process = items[start:end]
    print(f"Processing range [{start}, {end}), concurrency={args.concurrency}\n")

    # -------------------- Resume: load any existing output ------------------
    os.makedirs(os.path.dirname(os.path.abspath(args.output_file)) or ".", exist_ok=True)
    existing_by_key: Dict = {}
    existing_order: List = []
    if os.path.exists(args.output_file):
        try:
            with open(args.output_file, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        entry = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    k = make_key(entry)
                    if k not in existing_by_key:
                        existing_order.append(k)
                    existing_by_key[k] = entry  # newer overrides older
        except IOError:
            pass

    results = dict(existing_by_key)
    order = list(existing_order)
    order_seen = set(existing_order)

    total_items = end - start

    # -------------------- Build (entry, ranker) task list -------------------
    item_ctx = {}
    tasks = []          # [(idx, ranker), ...] in entry-major order
    direct_write = []   # entries that have nothing to score at all
    skip_write = []     # entries where every ranker is already done
    n_done_skip = n_full = n_repair = 0

    for i, it in enumerate(items_to_process):
        idx = start + i
        key = make_key(it)
        rubric_block = it.get(args.rubric_field) or {}
        rubrics_str, dim_rubrics = format_rubrics(rubric_block)
        n_rubrics = sum(len(v) for v in dim_rubrics.values())
        result_base = {k: v for k, v in it.items()}

        if n_rubrics == 0:
            if key not in existing_by_key:
                result_base["rubric_scores_by_ranker"] = {}
                result_base["scoring_error"] = "no rubric"
                direct_write.append((key, result_base))
            continue

        selected = it.get(args.selected_field) or {}
        avail = [r for r in rankers if isinstance(selected, dict) and r in selected]
        if not avail:
            if key not in existing_by_key:
                result_base["rubric_scores_by_ranker"] = {}
                result_base["scoring_error"] = "no ranker selected"
                direct_write.append((key, result_base))
            continue

        # Merge (a) rubric_scores_by_ranker that may already live in the
        # INPUT entry (so callers can provide partial scores and only ask us
        # to fill in the gaps) with (b) whatever the previous run wrote to
        # the OUTPUT file (which takes precedence).
        existing_blocks = {}
        if isinstance(it.get("rubric_scores_by_ranker"), dict):
            existing_blocks = dict(it["rubric_scores_by_ranker"])
        existing = existing_by_key.get(key)
        if isinstance(existing, dict) and isinstance(existing.get("rubric_scores_by_ranker"), dict):
            existing_blocks = {**existing_blocks, **existing["rubric_scores_by_ranker"]}

        by_ranker = {}
        ranker_tasks = {}   # ranker -> ("full"|"repair", missing_dims)
        for r in avail:
            work, missing = ranker_work(existing_blocks.get(r), dim_rubrics)
            if work == "done":
                by_ranker[r] = existing_blocks.get(r)
            else:
                ranker_tasks[r] = (work, missing)
                if work == "full":
                    n_full += 1
                else:
                    n_repair += 1

        if not ranker_tasks:
            n_done_skip += 1
            done_result = result_base
            done_result["rubric_scores_by_ranker"] = {
                r: by_ranker[r] for r in avail if r in by_ranker
            }
            done_result.pop("scoring_error", None)
            skip_write.append((key, done_result))
            continue

        item_ctx[idx] = {
            "item": it, "key": key, "result_base": result_base,
            "rubric_block": rubric_block, "rubrics_str": rubrics_str, "dim_rubrics": dim_rubrics,
            "avail": avail, "existing_blocks": existing_blocks,
            "by_ranker": by_ranker, "ranker_tasks": ranker_tasks,
            "remaining": len(ranker_tasks),
        }
        for r in avail:
            if r in ranker_tasks:
                tasks.append((idx, r))

    print(f"[Resume] {n_done_skip} entries fully done | pending: full={n_full}, repair={n_repair}")

    out_f = open(args.output_file, "a", encoding="utf-8")
    write_lock = threading.Lock()
    state_lock = threading.Lock()
    done_cnt, failed = 0, 0

    def _record(key, result):
        results[key] = result
        if key not in order_seen:
            order.append(key)
            order_seen.add(key)

    def _write_result(key, result):
        nonlocal done_cnt, failed
        line = json.dumps(result, ensure_ascii=False) + "\n"
        with write_lock:
            out_f.write(line)
            out_f.flush()
            try:
                os.fsync(out_f.fileno())
            except (OSError, ValueError):
                pass
            done_cnt += 1
            br = result.get("rubric_scores_by_ranker")
            if result.get("scoring_error") or not isinstance(br, dict) or len(br) == 0:
                failed += 1
            _record(key, result)

    for key, res in direct_write:
        _write_result(key, res)
    for key, res in skip_write:
        _write_result(key, res)
    if skip_write:
        print(f"[Resume] Flushed {len(skip_write)} already-complete entries to the output")

    def _handle_task(idx, ranker):
        ctx = item_ctx[idx]
        work, missing = ctx["ranker_tasks"][ranker]
        try:
            if work == "repair":
                block = repair_one_ranker(
                    idx, total_items, ctx["item"], ranker,
                    ctx["existing_blocks"].get(ranker) or {}, ctx["rubric_block"], missing,
                    args.docs_field, args.selected_field, args.max_docs,
                    prompt_template, args.app_id, args.app_key, args.base_url, args.model)
            else:  # full
                block = score_one_ranker(
                    idx, total_items, ctx["item"], ranker,
                    ctx["dim_rubrics"], ctx["rubrics_str"],
                    args.docs_field, args.selected_field, args.max_docs,
                    prompt_template, args.app_id, args.app_key, args.base_url, args.model)
        except Exception as e:
            tqdm.write(f"  [ERROR] entry {idx + 1} ranker={ranker} failed: {e}")
            if work == "repair":
                block = ctx["existing_blocks"].get(ranker) or {}
            else:
                empty = compute_scores({}, ctx["dim_rubrics"])
                block = {"scored_doc_ids": [], "rubric_scores": empty,
                         "overall": compute_overall(empty), "scoring_error": f"ERROR: {e}"}

        finalize = None
        with state_lock:
            ctx["by_ranker"][ranker] = block
            ctx["remaining"] -= 1
            if ctx["remaining"] == 0:
                result = ctx["result_base"]
                result["rubric_scores_by_ranker"] = {
                    r: ctx["by_ranker"][r] for r in ctx["avail"] if r in ctx["by_ranker"]
                }
                result.pop("scoring_error", None)
                finalize = result
        if finalize is not None:
            _write_result(ctx["key"], finalize)

    try:
        if args.concurrency == 1:
            for idx, ranker in tqdm(tasks, desc="rubric scoring", unit="task"):
                _handle_task(idx, ranker)
        else:
            with ThreadPoolExecutor(max_workers=args.concurrency) as executor:
                futures = [executor.submit(_handle_task, idx, ranker) for idx, ranker in tasks]
                for future in tqdm(as_completed(futures), total=len(futures),
                                   desc="rubric scoring", unit="task"):
                    future.result()
    finally:
        out_f.close()

    # Final rewrite: dedupe by key, using the latest result per key. This
    # also refreshes each block's `overall` so it always matches the current
    # dimension scores (after any late repair).
    tmp_path = args.output_file + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as wf:
        for k in order:
            if k in results:
                wf.write(json.dumps(refresh_entry_overall(results[k]), ensure_ascii=False) + "\n")
    os.replace(tmp_path, args.output_file)

    print(f"\nScoring finished. Results saved to: {args.output_file}")
    print(f"Written/updated this run: {done_cnt} | Total unique entries: {len(order)}")
    if failed:
        print(f"Failed/empty entries: {failed}")

    stats_path = write_stats(args.output_file, rankers)
    print(f"Excel summary saved to: {stats_path}")


if __name__ == "__main__":
    main()
